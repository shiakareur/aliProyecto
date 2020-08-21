from io import BytesIO
from flask_restful import Resource
from flask import request
from app import db
from sqlalchemy.exc import SQLAlchemyError
from models.Informacion import Informacion
from models.Modalidad import Modalidad
from models.Persona import Persona
from models.Promocion import Promocion
from openpyxl import load_workbook


class carga_masiva(Resource):
    def post(self):
        file = request.files['file']
        if not file:
            response = {'error': 'No input data provided'}
            return response
        else:
            workbook = load_workbook(filename=BytesIO(file.read()))
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, min_col=1, max_col=9, values_only=True):
                #print(row)
                obj_persona = Persona()
                contador=0
                for cell in row:
                    print(cell)
                    contador+=1
                    if contador==1:
                        obj_persona.codigo = cell
                    if contador==3:
                        obj_persona.nombres = cell
                    if contador==4:
                        obj_persona.apellido_paterno = cell
                    if contador==5:
                        obj_persona.apellido_materno = cell
                    if contador==6:
                        obj_persona.fecha_nacimiento = cell
                    if contador==7:
                        obj_persona.correo = cell
                    if contador==8:
                        obj_persona.telefono = cell
                    if contador==9:
                        obj_persona.direccion = cell
                obj_persona.add(obj_persona)
                db.session.commit()